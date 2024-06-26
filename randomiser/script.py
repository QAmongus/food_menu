import os
import time
import config
import random
import pandas as pd
from datetime import datetime
from prettytable import PrettyTable
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from signal import signal, alarm, SIGALRM
random.seed(int(time.time() * 1000))


def daily_menu(breakfast, lunch, dinner):
    print(f'Го {breakfast} на завтрак')
    print(f'На обед в {lunch}')
    print(f'Ужинаем в {dinner}')


def cooking():
    dish = random.choice(config.for_cooking)
    print(f'Готовим {dish}')


def chill():
    place_to_chill = random.choice(config.for_flex)
    print(f'Чиллим в {place_to_chill}')


while True:
    print("Сколько дней планируем? Введи любой вариант: день/неделя/готовим/чиллим")
    what_user_want = input().strip().lower()
    if what_user_want == "день":
        breakfast = random.choice(config.for_breakfast)
        lunch = random.choice(config.for_lunch)
        dinner = random.choice(config.for_dinner)

        while dinner == lunch:
            dinner = random.choice(config.for_dinner)

        daily_menu(breakfast, lunch, dinner)
        print("Все норм? Да/нет")
        all_is_ok = input().strip().lower()
        while all_is_ok != "да":
            print("Напиши завтрак/обед/ужин и я сгенерю новый")
            what_to_change = input().strip().lower()
            if what_to_change == "завтрак":
                new_breakfast = random.choice(config.for_breakfast)
                while new_breakfast == breakfast:
                    new_breakfast = random.choice(config.for_breakfast)
                breakfast = new_breakfast
                daily_menu(breakfast, lunch, dinner)
            elif what_to_change == "обед":
                new_lunch = random.choice(config.for_lunch)
                while new_lunch == lunch or new_lunch == dinner:
                    new_lunch = random.choice(config.for_lunch)
                lunch = new_lunch
                daily_menu(breakfast, lunch, dinner)
            elif what_to_change == "ужин":
                new_dinner = random.choice(config.for_dinner)
                while new_dinner == dinner or new_dinner == lunch:
                    new_dinner = random.choice(config.for_lunch)
                dinner = dinner
                daily_menu(breakfast, lunch, dinner)
            print("Все норм? Да/нет")
            all_is_ok = input().strip().lower()
            if all_is_ok == "да":
                break
            else:
                continue
    elif what_user_want == "неделя":
        table = PrettyTable()
        for day in config.day_names:
            breakfast = random.choice(config.for_breakfast)
            lunch = random.choice(config.for_lunch)
            dinner = random.choice(config.for_dinner)

            while dinner == lunch:
                dinner = random.choice(config.for_dinner)
            table.add_column(day, [breakfast, lunch, dinner])
        print(table)
        rows = [list(row) for row in table._rows]
        dataframe = pd.DataFrame(rows, columns=table.field_names)
        dataframe.to_excel("week_menu.xlsx", index=False, engine="openpyxl")

        workbook = load_workbook("week_menu.xlsx")
        worksheet = workbook.active
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            column_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[column_letter].width = max_length * 1.2

        workbook.save("week_menu.xlsx")
        print("Создал тебе week_menu.xlsx <3")
        # print("Отправить в телегу?")
        # user_option = input().strip().lower()
        # if user_option == "да":
        #     os.system("telegram-send --file week_menu.xlsx")
        #     print("Проверь табличку в тг")
        # else:
        #     continue
    elif what_user_want == "готовим":
        cooking()
    elif what_user_want == "чиллим":
        chill()
    else:
        print("ты дурак что ли?")
